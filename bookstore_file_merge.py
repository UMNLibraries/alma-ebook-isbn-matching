"""
This script produces a multiple-worksheet Excel .xlsx file of book metadata with proxied URLs derived from
four input files: a .csv file of Twin Cities-available ebooks with Portfolio IDs from Alma Analytics;
a .csv file of Portfolio IDs and proxied URLs from Alma's Export URLs job; a .pkl file to serve as a
concordance between Alma MMS IDs and Bookstore-supplied ISBNs; and an Excel spreadsheet supplied
by the Bookstore of course books for a given semester.
"""

import pandas as pd
import openpyxl
from datetime import date
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def add_isbns(ana_file):


	"""Adds Bookstore ISBNs to Alma Analytics available ebooks info based on MMS ID match; 
	returns the merged data as a pandas dataframe"""

	df_ana = pd.read_csv(ana_file)
	df_ana['Portfolio Id'] = df_ana['Portfolio Id'].astype(str)
	df_ana.rename(columns = {'Electronic Collection Public Name' : 'Supplier/Interface'}, inplace=True)
	df_ana['Publication Place'] = df_ana['Publication Place'].str.strip(':;')
	
	df_isbn = pd.read_pickle('isbns_found_all.pkl')

	df_isbn.columns = ['MMS Id', 'Bookstore ISBN']
	df_ana['MMS Id'] = df_ana['MMS Id'].astype(str)
	df_isbn['MMS Id'] = df_isbn['MMS Id'].astype(str)

	df_isbns = pd.merge(df_ana, df_isbn, on="MMS Id", how="left")
	df_isbns['MMS Id'] = df_isbns['MMS Id'].astype(str)

	return df_isbns

def isbn_url_merge(df_isbns, url_file):


	"""Merges the dataframe produced by add_isbns() with the data in the csv file produced by
	Alma's Export URLs job based on Portfolio ID match; returns the merged data as a pandas dataframe"""

	df_url = pd.read_csv(url_file)
	df_url.columns = ['Resource Type', 'Portfolio Id', 'URL']
	df_url['URL'] = df_url['URL'].str.replace('http://login.ezproxy','https://login.ezproxy')
	df_url['Portfolio Id'] = df_url['Portfolio Id'].astype(str)
	df_url['Portfolio Id'] = df_url['Portfolio Id'].str.strip("'")
	df_urls = pd.merge(df_isbns, df_url, on="Portfolio Id", how="left")

	return df_urls

	
def add_urls(df_urls, bkstore_file, bkstore_sheet):


	"""Merges the datframe produced by isbn_url_merge() with data from the Bookstore-supplied Excel
	spreadsheet on the basis of ISBN match; returns the merged data as a pandas dataframe."""

	df_urls_clean = df_urls.drop(['ISBN', 'Available For Group', 'Resource Type'], axis=1)
	df_urls_clean.rename(columns = {'Bookstore ISBN':'ISBN'}, inplace=True)
	df_urls_clean['ISBN'] = df_urls_clean['ISBN'].astype(str)
	
	df_bkstore = pd.ExcelFile(bkstore_file).parse(bkstore_sheet)
	df_bkstore['ISBN'] = df_bkstore['ISBN'].astype(str)

	df_all = pd.merge(df_bkstore, df_urls_clean, on="ISBN", how="left")
	df_all['ISBN'] = df_all['ISBN'].astype(str)

	return df_all


def make_spreadsheet(bkstore_file, df_urls, df_all):
	

	"""Appends several worksheets to the original Excel file, containing data extracted from Alma, and
	three views of the merged dataframe created by add_urls(): full list, books owned, and books not owned.""" 

	wb = load_workbook(bkstore_file)

	wb.create_sheet('Alma Data', 0)
	ws = wb['Alma Data']
	df_urls['Bookstore ISBN'] = df_urls['Bookstore ISBN'].astype(str)
	for r in dataframe_to_rows(df_urls, index=False, header=True):
		ws.append(r)

	wb.create_sheet('Full List', 0)
	ws = wb['Full List']
	df_all_clean = df_all.rename(columns = {'Title_x':'Title','Title_y':'Alma Title'})
	df_all_clean = df_all_clean.drop(['Portfolio Id', 'MMS Id'], axis=1)
	for r in dataframe_to_rows(df_all_clean, index=False, header=True):
		ws.append(r)

	wb.create_sheet('Already Owned', 0)
	ws = wb['Already Owned']
	df_owned = df_all[df_all['URL'].notnull()]
	df_owned = df_owned.drop(['Title_x', 'Portfolio Id', 'MMS Id'], axis=1)
	df_owned.rename(columns = {'Title_y':'Title'}, inplace=True)
	for r in dataframe_to_rows(df_owned, index=False, header=True):
		ws.append(r)

	wb.create_sheet('Not Owned', 0)
	ws = wb['Not Owned']
	df_not_owned = df_all[df_all['URL'].isnull()]
	df_not_owned = df_not_owned.drop(['Portfolio Id', 'MMS Id', 'Title_y', 'Supplier/Interface', 'Publisher', 'Publication Place', 
		'Publication Date', 'URL'], axis=1)
	df_not_owned.rename(columns = {'Title_x':'Title'}, inplace=True)
	for r in dataframe_to_rows(df_not_owned, index=False, header=True):
		ws.append(r)

	today = str(date.today())
	outfile = str(bkstore_file).rstrip('.xlsx')
	wb.save(outfile + '_mnu_ebooks_' + today + '.xlsx')


def main():


	"""Requests user input for filenames to process; calls add_isbns(), isbn_url_merge(),
	and add_urls(); writes two csv files: one of the final merge of Alma-derived data, and
	one of Bookstore spreadsheet data merged with Alma data."""

	ana_file = input("Alma Analytics csv filename: ")
	df_isbns = add_isbns(ana_file)

	url_file = input("Alma Export URL csv filename: ")
	df_urls = isbn_url_merge(df_isbns, url_file)

	bkstore_file = input("Bookstore .xlsx filename: ")
	bkstore_sheet = input("Bookstore sheet name: ")
	df_all = add_urls(df_urls, bkstore_file, bkstore_sheet)

	make_spreadsheet(bkstore_file, df_urls, df_all)

	#enable if needed for debugging
	#df_urls.to_csv('alma_data.csv')
	#df_all.to_csv('bkstore_merged.csv')

if __name__ == '__main__':
	main()